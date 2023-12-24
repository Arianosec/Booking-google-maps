from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl


def main():
    options = webdriver.ChromeOptions()
    options.add_argument(
        "User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install(), options=options))
    driver.get(
        "https://www.booking.com/searchresults.ru.html?ss=%D0%BF%D0%B0%D1%80%D0%B8%D0%B6&ssne=%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0&ssne_untouched=%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0&label=gog235jc-1DCAEoggI46AdIIVgDaE2IAQGYASG4ARjIAQzYAQPoAQH4AQKIAgGoAgS4ApHa0asGwAIB0gIkMTI3ZGU1YWYtOWVjNS00NzQ5LTgwNTQtN2VhMmUxZGM1ZjQ32AIE4AIB&aid=397594&lang=ru&sb=1&src_elem=sb&src=searchresults&dest_id=-1456928&dest_type=city&ac_position=0&ac_click_type=b&ac_langcode=ru&ac_suggestion_list_length=5&search_selected=true&search_pageview_id=18795f91617001ef&ac_meta=GhAxODc5NWY5MTYxNzAwMWVmIAAoATICcnU6CtC%2F0LDRgNC40LZAAEoAUAA%3D&group_adults=2&no_rooms=1&group_children=0")
    time.sleep(10)
    true_urls = []
    urls = driver.find_elements(By.XPATH,'//*[@id="bodyconstraint-inner"]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[3]/div/div[1]/div[2]/div/div/div/div[1]/div/div[1]/div/h3/a')
    for url in urls:
        true_urls.append(url.get_attribute('src'))

    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    index = 0
    # Создаем листы для различных сценариев
    sheet_google_booking = wb.create_sheet("Есть на гугл картах, есть ссылка на букинг.")
    sheet_google_no_booking = wb.create_sheet("Есть на гугл картах, нет ссылки на букинг")
    sheet_not_found = wb.create_sheet("Такого отеля нет")

    titles = driver.find_elements(By.XPATH,
                                  "/html/body/div[4]/div/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[3]/div/div[1]/div[2]/div/div/div/div[1]/div/div[1]/div/h3")

    for title in titles:
        hotel_name = title.text.replace("Откроется в новом окне", "").replace(" ", "+").replace("\n", "")
        print(hotel_name)

        driver.get(f"https://www.google.com/maps/search/{hotel_name}/@48.8731503,2.3157361,15z?entry=ttu")
        time.sleep(10)

        try:
            driver.find_element(By.XPATH,
                                "//*[@id='QA0Szd']/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[1]/h1")

            try:
                driver.find_element(By.XPATH, "//*[contains(text(),'Booking')]")
                sheet = sheet_google_booking
                print('Есть на гугл картах, есть ссылка на букинг.')
            except NoSuchElementException:
                sheet = sheet_google_no_booking
                print('Есть на гугл картах, нет ссылки на букинг')

        except NoSuchElementException:
            sheet = sheet_not_found
            print('Такого отеля нет')

        # Записываем данные в соответствующий лист
        sheet.append([hotel_name, true_urls[index]])
        index+=1

    # Сохраняем файл Excel
    wb.save("hotel_data.xlsx")


if __name__ == '__main__':
    main()
